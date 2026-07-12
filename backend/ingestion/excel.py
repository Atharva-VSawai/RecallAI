import openpyxl
import logging
from typing import List

logger = logging.getLogger(__name__)


def extract_text_from_excel(file_bytes: bytes, filename: str) -> str:
    """
    Extract text from Excel file in a compact format to avoid token limits.
    """
    try:
        from io import BytesIO
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        
        all_text_parts: List[str] = []
        sheet_names = workbook.sheetnames
        
        logger.info(f"[EXCEL] Processing '{filename}' with {len(sheet_names)} sheet(s)")
        
        total_rows = 0
        MAX_ROWS_PER_SHEET = 50  # Limit rows to avoid token overflow
        
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            
            if sheet.max_row == 0 or sheet.max_column == 0:
                continue
            
            all_text_parts.append(f"\n[SHEET: {sheet_name}]")
            
            # Read rows
            rows_data = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if idx >= MAX_ROWS_PER_SHEET + 1:  # +1 for header
                    break
                row_list = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(cell for cell in row_list):
                    rows_data.append(row_list)
            
            if not rows_data:
                continue
            
            # Headers
            headers = rows_data[0] if rows_data else []
            data_rows = rows_data[1:] if len(rows_data) > 1 else []
            
            if headers:
                all_text_parts.append("Columns: " + " | ".join(headers))
            
            # Data rows - compact format
            for row in data_rows:
                all_text_parts.append(" | ".join(row))
                total_rows += 1
            
            if sheet.max_row > MAX_ROWS_PER_SHEET + 1:
                all_text_parts.append(f"[Note: Showing first {MAX_ROWS_PER_SHEET} rows only]")
            
            logger.info(f"[EXCEL] Sheet '{sheet_name}': {len(data_rows)} rows extracted")
        
        full_text = "\n".join(all_text_parts)
        
        # Compact metadata
        metadata = f"Excel: {filename}\nSheets: {', '.join(sheet_names)}\nRows: {total_rows}\n"
        
        final_text = metadata + full_text
        logger.info(f"[EXCEL] Extracted {total_rows} rows, {len(final_text)} chars")
        
        return final_text
        
    except Exception as e:
        logger.error(f"[EXCEL] Error: {str(e)}")
        raise ValueError(f"Failed to extract Excel: {str(e)}")
